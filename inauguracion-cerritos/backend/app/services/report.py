"""Genera un reporte ejecutivo en Excel (.xlsx) con formato premium."""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..models import Lead, Prize, Spin

NAVY = "FF0A2E57"
NAVY_LIGHT = "FF12386B"
YELLOW = "FFFFD200"
GREEN = "FF1F9E5A"
RED = "FFE63329"
ORANGE = "FFB8860B"
WHITE = "FFFFFFFF"
LIGHT = "FFEEF3F8"
GREEN_SOFT = "FFE3F3EA"
RED_SOFT = "FFFBE7E6"

thin = Side(style="thin", color="FFD5DEE8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1,
                value=f"Tienda Pintuco Cerritos · Generado {datetime.utcnow():%Y-%m-%d %H:%M} UTC")
    s.font = Font(size=9, italic=True, color=NAVY_LIGHT)
    s.fill = PatternFill("solid", fgColor=LIGHT)
    ws.row_dimensions[2].height = 18


def _band(ws, row, text, ncols, color=NAVY_LIGHT):
    """Franja de sección dentro de una hoja."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, color=WHITE)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _header(ws, row, headers, color=NAVY_LIGHT):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color=WHITE, size=11)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 22


def _autosize(ws, widths):
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _kpi(ws, row, col, label, value, color):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(size=10, color=NAVY_LIGHT, bold=True)
    lc.alignment = Alignment(horizontal="center")
    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = Font(size=20, bold=True, color=WHITE)
    vc.fill = PatternFill("solid", fgColor=color)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = BORDER
    ws.row_dimensions[row + 1].height = 38


def _si_no(ws, cell, es_si):
    cell.alignment = Alignment(horizontal="center")
    if es_si:
        cell.value = "SÍ"
        cell.font = Font(bold=True, color=GREEN)
        cell.fill = PatternFill("solid", fgColor=GREEN_SOFT)
    else:
        cell.value = "Pendiente"
        cell.font = Font(bold=True, color=RED)
        cell.fill = PatternFill("solid", fgColor=RED_SOFT)


def _fecha(dt):
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


def generar_reporte(db: Session) -> bytes:
    wb = Workbook()

    # Métricas base
    total_part = db.query(func.count(Lead.id)).scalar() or 0
    total_giros = db.query(func.count(Spin.id)).scalar() or 0
    ganados = db.query(func.count(Spin.id)).filter(Spin.gano.is_(True)).scalar() or 0
    prem_entregados = db.query(func.count(Spin.id)).filter(Spin.redeemed.is_(True)).scalar() or 0
    disponibles = db.query(func.coalesce(func.sum(Prize.stock_restante), 0)).filter(
        Prize.activo.is_(True), Prize.es_perdedor.is_(False)).scalar() or 0
    conv = round((total_giros / total_part * 100), 1) if total_part else 0.0
    referidos = db.query(func.count(Lead.id)).filter(Lead.referred_by.isnot(None)).scalar() or 0
    cup_usados = db.query(func.count(Lead.id)).filter(Lead.coupon_redeemed.is_(True)).scalar() or 0
    cup_pend = total_part - cup_usados
    prem_pend = ganados - prem_entregados

    # ---------- Hoja 1: Resumen ----------
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    _title(ws, "📊 REPORTE EJECUTIVO — INAUGURACIÓN", 6)

    _kpi(ws, 4, 1, "PARTICIPANTES", total_part, NAVY)
    _kpi(ws, 4, 2, "GIROS", total_giros, NAVY_LIGHT)
    _kpi(ws, 4, 3, "CONVERSIÓN", f"{conv}%", GREEN)
    _kpi(ws, 4, 4, "PREMIOS GANADOS", ganados, RED)

    _band(ws, 7, "🏆 PREMIOS", 4, NAVY)
    _kpi(ws, 8, 1, "GANADOS", ganados, NAVY_LIGHT)
    _kpi(ws, 8, 2, "ENTREGADOS", prem_entregados, GREEN)
    _kpi(ws, 8, 3, "POR ENTREGAR", prem_pend, ORANGE)
    _kpi(ws, 8, 4, "DISPONIBLES (stock)", int(disponibles), NAVY)

    _band(ws, 11, "🎁 CUPONES 10% DE DESCUENTO", 4, NAVY)
    _kpi(ws, 12, 1, "EMITIDOS", total_part, NAVY_LIGHT)
    _kpi(ws, 12, 2, "USADOS", cup_usados, GREEN)
    _kpi(ws, 12, 3, "POR USAR", cup_pend, ORANGE)
    _kpi(ws, 12, 4, "REFERIDOS", referidos, NAVY)

    # Desglose por premio
    _band(ws, 15, "Desglose por premio", 5)
    _header(ws, 16, ["Premio", "Stock total", "Restante", "Ganados", "Entregados"])
    prizes = db.query(Prize).order_by(Prize.orden.asc()).all()
    r = 17
    for p in prizes:
        pg = db.query(func.count(Spin.id)).filter(Spin.prize_id == p.id).scalar() or 0
        pe = db.query(func.count(Spin.id)).filter(
            Spin.prize_id == p.id, Spin.redeemed.is_(True)).scalar() or 0
        for j, v in enumerate([p.nombre, p.stock_total, p.stock_restante, pg, pe], start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j > 1 else "left")
        r += 1
    _autosize(ws, [24, 14, 14, 14, 16])

    # ---------- Hoja 2: Participantes (con estado del cupón) ----------
    ws2 = wb.create_sheet("Participantes")
    ws2.sheet_view.showGridLines = False
    _title(ws2, "👥 PARTICIPANTES INSCRITOS", 10)
    _header(ws2, 4, ["Nombre", "Teléfono", "Correo", "Cédula", "Dirección",
                     "Cupón", "¿Bono usado?", "Fecha bono", "Referido por", "Fecha registro"])
    leads = db.query(Lead).order_by(Lead.created_at.desc()).all()
    for i, l in enumerate(leads):
        row = 5 + i
        vals = [l.nombre, l.telefono, l.correo, l.cedula, l.direccion or "",
                l.coupon_code, None, _fecha(l.coupon_redeemed_at),
                l.referred_by or "", _fecha(l.created_at)]
        for j, v in enumerate(vals, start=1):
            c = ws2.cell(row=row, column=j, value=v)
            c.border = BORDER
            if row % 2 == 0 and j != 7:
                c.fill = PatternFill("solid", fgColor=LIGHT)
        _si_no(ws2, ws2.cell(row=row, column=7), l.coupon_redeemed)
    ws2.freeze_panes = "A5"
    _autosize(ws2, [24, 15, 28, 15, 22, 17, 13, 17, 15, 17])

    # ---------- Hoja 3: Premios (giros ganadores + estado entrega) ----------
    ws3 = wb.create_sheet("Premios")
    ws3.sheet_view.showGridLines = False
    _title(ws3, "🏆 PREMIOS GANADOS Y ENTREGA", 9)
    _header(ws3, 4, ["Cliente", "Teléfono", "Cédula", "Premio", "¿Entregado?",
                     "Entregado por", "Fecha giro", "Fecha entrega", "Correo"])
    ganadores = db.query(Spin).filter(Spin.gano.is_(True)).order_by(Spin.created_at.desc()).all()
    for i, s in enumerate(ganadores):
        row = 5 + i
        lead = db.query(Lead).filter(Lead.id == s.lead_id).first()
        prize = db.query(Prize).filter(Prize.id == s.prize_id).first() if s.prize_id else None
        vals = [lead.nombre if lead else "", lead.telefono if lead else "",
                lead.cedula if lead else "", prize.nombre if prize else "Premio",
                None, s.redeemed_by or "", _fecha(s.created_at), _fecha(s.redeemed_at),
                lead.correo if lead else ""]
        for j, v in enumerate(vals, start=1):
            c = ws3.cell(row=row, column=j, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if j in (1, 4, 6, 9) else "center")
        _si_no(ws3, ws3.cell(row=row, column=5), s.redeemed)
    ws3.freeze_panes = "A5"
    _autosize(ws3, [24, 15, 15, 22, 13, 20, 18, 18, 26])

    # ---------- Hoja 4: Por Reclamar (pendientes) ----------
    ws4 = wb.create_sheet("Por Reclamar")
    ws4.sheet_view.showGridLines = False
    _title(ws4, "⏳ PENDIENTES POR RECLAMAR", 5)

    # A) Premios ganados NO entregados
    _band(ws4, 4, f"🏆 PREMIOS POR ENTREGAR ({prem_pend})", 5, ORANGE)
    _header(ws4, 5, ["Cliente", "Teléfono", "Correo", "Premio", "Fecha giro"])
    pend_prem = db.query(Spin).filter(
        Spin.gano.is_(True), Spin.redeemed.is_(False)).order_by(Spin.created_at.desc()).all()
    r = 6
    for s in pend_prem:
        lead = db.query(Lead).filter(Lead.id == s.lead_id).first()
        prize = db.query(Prize).filter(Prize.id == s.prize_id).first() if s.prize_id else None
        vals = [lead.nombre if lead else "", lead.telefono if lead else "",
                lead.correo if lead else "", prize.nombre if prize else "Premio",
                _fecha(s.created_at)]
        for j, v in enumerate(vals, start=1):
            c = ws4.cell(row=r, column=j, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if j != 5 else "center")
        r += 1
    if not pend_prem:
        ws4.cell(row=r, column=1, value="— Sin premios pendientes —").font = Font(italic=True, color=NAVY_LIGHT)
        r += 1

    # B) Cupones NO usados
    r += 1
    _band(ws4, r, f"🎁 CUPONES 10% POR USAR ({cup_pend})", 5, ORANGE)
    r += 1
    _header(ws4, r, ["Cliente", "Teléfono", "Correo", "Cupón", "Fecha registro"])
    r += 1
    pend_cup = db.query(Lead).filter(Lead.coupon_redeemed.is_(False)).order_by(
        Lead.created_at.desc()).all()
    for l in pend_cup:
        vals = [l.nombre, l.telefono, l.correo, l.coupon_code, _fecha(l.created_at)]
        for j, v in enumerate(vals, start=1):
            c = ws4.cell(row=r, column=j, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if j != 5 else "center")
        r += 1
    if not pend_cup:
        ws4.cell(row=r, column=1, value="— Sin cupones pendientes —").font = Font(italic=True, color=NAVY_LIGHT)
    _autosize(ws4, [24, 15, 28, 20, 18])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
